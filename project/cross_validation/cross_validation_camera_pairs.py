#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
File: /workspace/Skiing_Canonical_DualView_3D_Pose_PyTorch/project/cross_validation_camera_pairs.py
Project: /workspace/Skiing_Canonical_DualView_3D_Pose_PyTorch/project
Created Date: Sunday March 9th 2026
Author: Kaixu Chen
-----
Comment:
交叉验证脚本 - 用于摄像头两两组合的场景
针对2个人物、12个动作、每个动作108个摄像头的数据集。
支持三种划分策略：
1. by_person: 按人物划分（Leave-One-Person-Out）
2. by_action: 按动作划分（K-Fold on actions）
3. by_camera_pair: 按摄像头对划分（K-Fold on camera pairs）

Have a good code time :)
-----
Copyright (c) 2026 The University of Tsukuba
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict, fields
from itertools import combinations

import numpy as np
from sklearn.model_selection import KFold

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


@dataclass
class CameraPairSample:
    """表示一个训练样本，按照人物的turn进行划分"""

    person_id: str  # "01", "02"
    turn_id: Optional[str] = None

    # video frames
    cam1_video_path: Optional[str] = None
    cam2_video_path: Optional[str] = None
    # sam 3d body path
    sam3d_cam1_results_path: Optional[str] = None
    sam3d_cam2_results_path: Optional[str] = None

    cam1_turn_frame_start: Optional[int] = None
    cam1_turn_frame_end: Optional[int] = None
    cam2_turn_frame_start: Optional[int] = None
    cam2_turn_frame_end: Optional[int] = None

    label_twist_3class: Optional[int] = None
    label_posture_3class: Optional[int] = None
    label_relax_3class: Optional[int] = None
    label_total_3class: Optional[int] = None

    fused_kpt_path: Optional[str] = None  # 预先计算的融合后的关键点文件路径

    def to_dict(self):
        return asdict(self)

    @classmethod
    def from_dict(cls, d: Dict):
        known = {f.name for f in fields(cls)}
        return cls(**{k: v for k, v in d.items() if k in known})


class CameraPairCrossValidation:
    """
    针对摄像头对的交叉验证策略

    参数:
        data_root: 数据根目录
        num_persons: 人物数量（默认2）
        num_actions: 动作数量（默认12）
        num_cameras: 每个动作的摄像头数量（默认108）
        split_strategy: 划分策略，可选 'by_person', 'by_action', 'by_camera_pair'
        n_splits: K折交叉验证的折数（仅用于 by_action 和 by_camera_pair 策略）
        index_save_path: 索引文件保存路径
    """

    def __init__(
        self,
        data_root: str,
        num_persons: int = 2,
        num_actions: int = 12,
        num_cameras: int = 108,
        split_strategy: str = "by_person",  # by_person, by_action, by_camera_pair
        n_splits: int = 5,
        sam3d_export_root: Optional[str] = None,
        annotation_path: Optional[str] = None,
        fused_kpt_root: Optional[str] = None,
        index_save_path: Optional[str] = None,
    ):
        self.data_root = Path(data_root)
        self.data_dir = self.data_root / "data"
        if sam3d_export_root and str(sam3d_export_root).strip():
            self.sam3d_export_root = Path(sam3d_export_root)
        else:
            raw_sam3d_person = self.data_root / "sam3d_body_results" / "person"
            self.sam3d_export_root = (
                raw_sam3d_person
                if raw_sam3d_person.exists()
                else self.data_root / "modalities_from_sam3d"
            )
        if annotation_path and str(annotation_path).strip():
            self.annotation_path = Path(annotation_path)
        else:
            default_annotation = self.data_root / "suwabe_label.xlsx"
            self.annotation_path = (
                default_annotation if default_annotation.exists() else None
            )
        if fused_kpt_root and str(fused_kpt_root).strip():
            self.fused_kpt_root = Path(fused_kpt_root)
        else:
            self.fused_kpt_root = self.data_root / "fuse"
        self.num_persons = num_persons
        self.num_actions = num_actions
        self.num_cameras = num_cameras
        self.split_strategy = split_strategy
        self.n_splits = n_splits

        if index_save_path is None:
            self.index_save_path: Path = (
                self.data_root / "index_mapping" / f"camera_pairs_{split_strategy}.json"
            )
        else:
            self.index_save_path = Path(index_save_path)
        self.index_save_path = self.index_save_path.resolve()

        self.index_save_path.parent.mkdir(parents=True, exist_ok=True)
        self._person_label_map: Dict[str, Dict[str, int]] = (
            self._load_annotation_labels()
        )

    @staticmethod
    def _capture_to_kpt2d_id(capture_name: str) -> str:
        # frames目录是 capture_Lx_Ayyy；kpt2d目录是 Lx_Ayyy
        return capture_name.replace("capture_", "", 1)

    def _discover_people_actions(self) -> Dict[str, List[str]]:
        if not self.data_dir.exists():
            raise FileNotFoundError(f"data目录不存在: {self.data_dir}")

        people_actions: Dict[str, List[str]] = {}
        for person_dir in sorted(p for p in self.data_dir.iterdir() if p.is_dir()):
            # 跳过辅助目录
            if person_dir.name.lower() in {"logs", "cameras"}:
                continue
            action_names: List[str] = []
            for action_dir in sorted(p for p in person_dir.iterdir() if p.is_dir()):
                if (action_dir / "frames").exists():
                    action_names.append(action_dir.name)
            if action_names:
                people_actions[person_dir.name] = action_names
        return people_actions

    @staticmethod
    def _normalize_three_class(value: Any) -> int:
        if value is None:
            raise ValueError("Empty label value in annotation")
        v = float(value)

        if np.isnan(v):
            raise ValueError("NaN label value in annotation")

        if 1.0 <= v <= 3.0:
            return int(round(v)) - 1

        if 0.0 <= v <= 2.0:
            return int(round(v))

        if v <= 1.5:
            return 0
        if v <= 2.5:
            return 1
        return 2

    def _load_annotation_labels(self) -> Dict[str, Dict[str, int]]:
        if self.annotation_path is None:
            return {}
        if not self.annotation_path.exists():
            raise FileNotFoundError(f"annotation 文件不存在: {self.annotation_path}")
        if load_workbook is None:
            raise ImportError(
                "openpyxl is required to read annotation xlsx. Please install openpyxl."
            )

        wb = load_workbook(self.annotation_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_map = {name: idx for idx, name in enumerate(header)}

        def find_col(candidates: List[str]) -> Optional[int]:
            for c in candidates:
                if c in col_map:
                    return col_map[c]
            for k, idx in col_map.items():
                for c in candidates:
                    if c and c in k:
                        return idx
            return None

        id_idx = find_col(["ID", "id", "person_id", "人物ID"])
        twist_idx = find_col(["ねじり", "twist"])
        posture_idx = find_col(["姿勢", "posture"])
        relax_idx = find_col(["脱力", "relax"])

        if (
            id_idx is None
            or twist_idx is None
            or posture_idx is None
            or relax_idx is None
        ):
            raise KeyError(
                "annotation 表头缺少必要列，至少需要: ID, ねじり, 姿勢, 脱力"
            )

        person_map: Dict[str, Dict[str, int]] = {}
        for row in rows[1:]:
            if row is None:
                continue
            rid = row[id_idx] if id_idx < len(row) else None
            if rid is None:
                continue
            person_id = str(rid).strip()
            if person_id.endswith(".0"):
                person_id = person_id[:-2]
            person_id = person_id.lstrip("0") or "0"

            twist_raw = row[twist_idx] if twist_idx < len(row) else None
            posture_raw = row[posture_idx] if posture_idx < len(row) else None
            relax_raw = row[relax_idx] if relax_idx < len(row) else None

            twist = self._normalize_three_class(twist_raw)
            posture = self._normalize_three_class(posture_raw)
            relax = self._normalize_three_class(relax_raw)

            total_avg = (twist + posture + relax) / 3.0
            total = int(np.clip(int(round(total_avg)), 0, 2))

            person_map[person_id] = {
                "label_twist_3class": int(twist),
                "label_posture_3class": int(posture),
                "label_relax_3class": int(relax),
                "label_total_3class": int(total),
            }

        print(
            f"✓ 从 annotation 读取到 {len(person_map)} 个人物标签: {self.annotation_path}"
        )
        return person_map

    def _is_turn_dataset_layout(self) -> bool:
        return (self.data_root / "raw" / "person").exists() and (
            self.data_root / "split_cycle"
        ).exists()

    @staticmethod
    def _person_sort_key(person_id: str) -> Tuple[int, str]:
        token = str(person_id)
        digits = "".join(ch for ch in token if ch.isdigit())
        if digits:
            return int(digits), token
        return 10**9, token

    @staticmethod
    def _extract_turn_index(turn_id: str) -> int:
        digits = "".join(ch for ch in str(turn_id) if ch.isdigit())
        if digits:
            return int(digits)
        return 10**9

    @staticmethod
    def _pick_cycle_video(
        cycle_dir: Path, view: str, cycle_index: int
    ) -> Optional[Path]:
        view_dir = cycle_dir / view
        if not view_dir.exists():
            return None

        exact_prefix = f"cycle_{cycle_index:03d}_"
        exact_matches = sorted(view_dir.glob(f"{exact_prefix}*.mp4"))
        if exact_matches:
            return exact_matches[0]

        generic_matches = sorted(view_dir.glob("cycle_*.mp4"))
        if cycle_index < len(generic_matches):
            return generic_matches[cycle_index]
        return None

    def _discover_turn_samples(self) -> List[CameraPairSample]:
        raw_person_root = self.data_root / "raw" / "person"
        split_cycle_root = self.data_root / "split_cycle"
        sam3d_person_root = self.sam3d_export_root

        samples: List[CameraPairSample] = []

        person_dirs = sorted(
            (p for p in raw_person_root.iterdir() if p.is_dir()),
            key=lambda p: self._person_sort_key(p.name),
        )

        for person_dir in person_dirs:
            person_id = str(person_dir.name)
            cycle_dir = split_cycle_root / f"person_{person_id}"
            if not cycle_dir.exists():
                continue

            align_files = sorted(cycle_dir.glob("alignment_record_*.json"))
            if not align_files:
                continue
            align_path = align_files[0]

            with open(align_path, "r", encoding="utf-8") as f:
                align_record = json.load(f)
            cycles = align_record.get("cycles", [])
            if not isinstance(cycles, list):
                continue

            sam3d_face_dir = sam3d_person_root / person_id / "face"
            sam3d_side_dir = sam3d_person_root / person_id / "side"
            if person_id not in self._person_label_map:
                raise KeyError(
                    f"人物 {person_id} 在 annotation 中找不到标签: {self.annotation_path}"
                )
            person_label = self._person_label_map[person_id]
            fused_kpt_path = self.fused_kpt_root / f"person_{person_id}" / "frames"

            for cycle in cycles:
                if not isinstance(cycle, dict):
                    continue

                cycle_index = int(cycle.get("cycle_index", -1))
                if cycle_index < 0:
                    continue
                turn_id = f"turn_{cycle_index:03d}"

                face_video = self._pick_cycle_video(cycle_dir, "face", cycle_index)
                side_video = self._pick_cycle_video(cycle_dir, "side", cycle_index)
                if face_video is None or side_video is None:
                    continue

                face_frames = cycle.get("face_video_frames", {})
                side_frames = cycle.get("side_video_frames", {})

                sample = CameraPairSample(
                    person_id=person_id,
                    turn_id=turn_id,
                    cam1_video_path=str(face_video.resolve()),
                    cam2_video_path=str(side_video.resolve()),
                    sam3d_cam1_results_path=str(sam3d_face_dir.resolve()),
                    sam3d_cam2_results_path=str(sam3d_side_dir.resolve()),
                    cam1_turn_frame_start=int(face_frames.get("start", 0)),
                    cam1_turn_frame_end=int(face_frames.get("end", 0)),
                    cam2_turn_frame_start=int(side_frames.get("start", 0)),
                    cam2_turn_frame_end=int(side_frames.get("end", 0)),
                    label_twist_3class=int(person_label["label_twist_3class"]),
                    label_posture_3class=int(person_label["label_posture_3class"]),
                    label_relax_3class=int(person_label["label_relax_3class"]),
                    label_total_3class=int(person_label["label_total_3class"]),
                    fused_kpt_path=str(fused_kpt_path),
                )
                samples.append(sample)

        samples.sort(
            key=lambda s: (
                self._person_sort_key(s.person_id),
                self._extract_turn_index(s.turn_id or ""),
            )
        )

        print(f"✓ 总共生成 {len(samples)} 个 turn 样本")
        print(f"  - 人物数: {len({s.person_id for s in samples})}")
        print(f"  - turn数: {len({s.turn_id for s in samples if s.turn_id})}")
        print("  - 每个样本包含: 双视角视频路径 + 双视角SAM3D结果路径")

        return samples

    def build_all_samples(self) -> List[CameraPairSample]:
        """
        扫描真实目录，构建样本：person × action × camera_capture_pairs

        Returns:
            所有样本的列表
        """
        if self._is_turn_dataset_layout():
            return self._discover_turn_samples()

        samples: List[CameraPairSample] = []
        people_actions = self._discover_people_actions()

        action_count_total = 0
        per_action_pair_count: List[int] = []

        for person_id, actions in people_actions.items():
            for action_id in actions:
                action_count_total += 1
                action_dir = self.data_dir / person_id / action_id
                frames_root = action_dir / "frames"
                kpt2d_root = action_dir / "kpt2d"
                kpt3d_dir = action_dir / "kpt3d"
                meta_dir = action_dir / "meta"
                sam3d_export_action = self.sam3d_export_root / person_id / action_id

                capture_dirs = sorted(
                    p
                    for p in frames_root.iterdir()
                    if p.is_dir() and p.name.startswith("capture_")
                )
                if len(capture_dirs) < 2:
                    continue

                per_action_pair_count.append(
                    len(capture_dirs) * (len(capture_dirs) - 1) // 2
                )

                for cam1_dir, cam2_dir in combinations(capture_dirs, 2):
                    cam1_id = cam1_dir.name
                    cam2_id = cam2_dir.name

                    kpt2d_cam1 = kpt2d_root / self._capture_to_kpt2d_id(cam1_id)
                    kpt2d_cam2 = kpt2d_root / self._capture_to_kpt2d_id(cam2_id)

                    sam3d_cam1_kpt2d = sam3d_export_action / "kpt2d" / cam1_id
                    sam3d_cam2_kpt2d = sam3d_export_action / "kpt2d" / cam2_id
                    sam3d_cam1_kpt3d = sam3d_export_action / "kpt3d" / cam1_id
                    sam3d_cam2_kpt3d = sam3d_export_action / "kpt3d" / cam2_id

                    sequence_meta = meta_dir / "sequence.json"
                    joint_meta = meta_dir / "joint_names.json"

                    sample = CameraPairSample(
                        person_id=person_id,
                        turn_id=f"{action_id}_{cam1_id}_{cam2_id}",
                        cam1_video_path=str(cam1_dir.resolve()),
                        cam2_video_path=str(cam2_dir.resolve()),
                        sam3d_cam1_results_path=str(sam3d_export_action.resolve()),
                        sam3d_cam2_results_path=str(sam3d_export_action.resolve()),
                    )
                    samples.append(sample)

        people_count = len(people_actions)
        action_count = sum(len(v) for v in people_actions.values())
        avg_pairs = int(np.mean(per_action_pair_count)) if per_action_pair_count else 0

        print(f"✓ 总共生成 {len(samples)} 个样本")
        print(f"  - {people_count} 个人物")
        print(f"  - {action_count} 个动作")
        print(f"  - 每个动作平均 {avg_pairs} 个摄像头对")

        return samples

    def _sort_samples_by_person_turn(
        self, samples: List[CameraPairSample]
    ) -> List[CameraPairSample]:
        return sorted(
            samples,
            key=lambda s: (
                self._person_sort_key(s.person_id),
                self._extract_turn_index(s.turn_id or ""),
            ),
        )

    def split_by_turn(
        self, samples: List[CameraPairSample]
    ) -> Dict[int, Dict[str, Any]]:
        """
        按人物划分，同一个人的所有 turn 应该只属于 train、val 或 test 中的一个。
        策略：先按人物划分，每个人物的所有 turn 分配到同一个集合。
        """
        fold_dict: Dict[int, Dict[str, Any]] = {}

        # 按人物分组
        person_samples: Dict[str, List[CameraPairSample]] = {}
        for sample in samples:
            if sample.person_id not in person_samples:
                person_samples[sample.person_id] = []
            person_samples[sample.person_id].append(sample)

        if not person_samples:
            return {0: {"train": [], "val": [], "test": [], "split_by": "person_turn"}}

        rng = np.random.default_rng(42)
        n_splits = max(1, int(self.n_splits))

        # 获取所有人物 ID，按排序规则排序
        person_ids = sorted(person_samples.keys(), key=self._person_sort_key)

        for fold_idx in range(n_splits):
            train_samples = []
            val_samples = []
            test_samples = []

            # 对人物列表进行随机划分
            shuffled_persons = list(person_ids)
            rng_fold = np.random.default_rng(
                int(rng.integers(0, 10_000_000)) + fold_idx
            )
            rng_fold.shuffle(shuffled_persons)

            n_total_persons = len(shuffled_persons)
            n_train_persons = int(round(n_total_persons * 0.7))
            n_val_persons = int(round(n_total_persons * 0.2))
            n_test_persons = n_total_persons - n_train_persons - n_val_persons

            # 保证至少有一个人在 train/val/test 中
            if n_total_persons >= 3:
                if n_train_persons <= 0:
                    n_train_persons = 1
                if n_val_persons <= 0:
                    n_val_persons = 1
                n_test_persons = n_total_persons - n_train_persons - n_val_persons
                if n_test_persons <= 0:
                    n_test_persons = 1
                    if n_train_persons > n_val_persons:
                        n_train_persons -= 1
                    else:
                        n_val_persons -= 1

            train_person_ids = set(shuffled_persons[:n_train_persons])
            val_person_ids = set(
                shuffled_persons[n_train_persons : n_train_persons + n_val_persons]
            )
            test_person_ids = set(shuffled_persons[n_train_persons + n_val_persons :])

            # 根据人物 ID 分配所有样本
            for person_id in person_ids:
                person_all_samples = person_samples[person_id]

                if person_id in train_person_ids:
                    train_samples.extend(person_all_samples)
                elif person_id in val_person_ids:
                    val_samples.extend(person_all_samples)
                elif person_id in test_person_ids:
                    test_samples.extend(person_all_samples)

            # 排序样本
            train_samples = self._sort_samples_by_person_turn(train_samples)
            val_samples = self._sort_samples_by_person_turn(val_samples)
            test_samples = self._sort_samples_by_person_turn(test_samples)

            fold_dict[fold_idx] = {
                "train": train_samples,
                "val": val_samples,
                "test": test_samples,
                "split_by": "person",
                "train_persons": sorted(list(train_person_ids)),
                "val_persons": sorted(list(val_person_ids)),
                "test_persons": sorted(list(test_person_ids)),
                "sort_by": "person_id",
                "ratio": "7/2/1",
            }

            print(
                f"Fold {fold_idx}: train={len(train_samples)}, "
                f"val={len(val_samples)}, test={len(test_samples)} "
                f"(persons: train={n_train_persons}, val={n_val_persons}, test={n_test_persons})"
            )

        return fold_dict

    def prepare_folds(self) -> Dict[int, Dict[str, Any]]:
        """
        根据选择的策略准备交叉验证的折
        按人物划分：同一个人的所有 turn 只属于 train、val 或 test 中的一个
        """
        print(f"\n{'=' * 60}")
        print("准备交叉验证数据集")
        print(f"{'=' * 60}")
        print(f"策略: {self.split_strategy}")
        print(f"数据根目录: {self.data_root}")
        print(f"{'=' * 60}\n")

        samples = self.build_all_samples()
        fold_dict = self.split_by_turn(samples)

        return fold_dict

    def save_folds(self, fold_dict: Dict[int, Dict[str, Any]]):
        """
        保存交叉验证的划分结果到JSON文件
        """
        # 序列化
        serialized: Dict[str, Any] = {}
        for fold_idx, fold_data in fold_dict.items():
            serialized[str(fold_idx)] = {
                "train": [s.to_dict() for s in fold_data["train"]],
                "val": [s.to_dict() for s in fold_data["val"]],
                "test": [s.to_dict() for s in fold_data.get("test", [])],
            }
            # 保存额外信息（如验证集的人物或动作）
            for key in fold_data:
                if key not in ["train", "val", "test"]:
                    serialized[str(fold_idx)][key] = fold_data[key]

        # 添加元数据
        serialized["_metadata"] = {
            "num_persons": self.num_persons,
            "num_actions": self.num_actions,
            "num_cameras": self.num_cameras,
            "split_strategy": self.split_strategy,
            "n_splits": len(fold_dict),
            "total_samples": sum(
                len(fold_data["train"])
                + len(fold_data["val"])
                + len(fold_data.get("test", []))
                for fold_data in fold_dict.values()
            )
            // len(fold_dict),
        }

        with open(self.index_save_path, "w", encoding="utf-8") as f:
            json.dump(serialized, f, ensure_ascii=False, indent=2)

        print(f"\n✓ 交叉验证索引已保存到: {self.index_save_path}")

    def load_folds(self) -> Dict[int, Dict[str, Any]]:
        """
        从JSON文件加载交叉验证的划分结果
        """
        if not self.index_save_path.exists():
            raise FileNotFoundError(f"索引文件不存在: {self.index_save_path}")

        with open(self.index_save_path, "r", encoding="utf-8") as f:
            serialized = json.load(f)

        # 提取元数据
        metadata = serialized.pop("_metadata", {})
        print("\n加载交叉验证数据:")
        print(f"  策略: {metadata.get('split_strategy', 'unknown')}")
        print(f"  折数: {metadata.get('n_splits', 'unknown')}")
        print(f"  总样本数: {metadata.get('total_samples', 'unknown')}")

        # 反序列化
        fold_dict: Dict[int, Dict[str, Any]] = {}
        for fold_idx_str, fold_data in serialized.items():
            fold_idx = int(fold_idx_str)
            fold_dict[fold_idx] = {
                "train": [CameraPairSample.from_dict(d) for d in fold_data["train"]],
                "val": [CameraPairSample.from_dict(d) for d in fold_data["val"]],
                "test": [
                    CameraPairSample.from_dict(d) for d in fold_data.get("test", [])
                ],
            }
            # 恢复额外信息
            for key in fold_data:
                if key not in ["train", "val", "test"]:
                    fold_dict[fold_idx][key] = fold_data[key]

        return fold_dict

    def __call__(self, force_recreate: bool = False) -> Dict[int, Dict[str, Any]]:
        """
        主入口：创建或加载交叉验证划分

        Args:
            force_recreate: 是否强制重新创建索引文件
        """
        if self.index_save_path.exists() and not force_recreate:
            print("✓ 发现已存在的索引文件，直接加载")
            return self.load_folds()
        else:
            print("✓ 创建新的交叉验证划分")
            fold_dict = self.prepare_folds()
            self.save_folds(fold_dict)
            return fold_dict


def main():
    """
    示例使用
    """
    for strategy in ["by_person", "by_action", "by_camera_pair"]:
        cv = CameraPairCrossValidation(
            data_root="/workspace/data/skiing_unity_dataset",
            split_strategy=strategy,
            n_splits=5,
            index_save_path=f"/workspace/data/skiing_unity_dataset/index_mapping/camera_pairs_{strategy}.json",
        )
        folds = cv(force_recreate=True)
        print(f"\n示例 - {strategy}:")
        print(f"折数: {len(folds)}")


if __name__ == "__main__":
    main()
